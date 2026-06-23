#!/usr/bin/env python3
"""
Generate advisor-review.xlsx — a two-sheet Excel workbook for human review
of the scraped program JSON files before publication.

Usage:
    python3 scripts/generate-review-workbook.py
    python3 scripts/generate-review-workbook.py --out path/to/output.xlsx

Output: output/review/advisor-review.xlsx (or --out path)
"""

import argparse
import glob
import json
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_URL = "https://aei-web.apps.paas02-t.ilstu.edu/advising/html/"
PROGRAMS_DIR = "data/programs"

# ── colours ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="C00000")   # ISU red
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")
NORMAL_FONT   = Font(size=10)
LINK_FONT     = Font(size=10, color="0563C1", underline="single")
BOLD_FONT     = Font(bold=True, size=10)
THIN_BORDER   = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)

# Issue types used in Sheet 2
ISSUE_AUTO     = "auto_fulfilled_by"
ISSUE_CREDIT   = "credit_assumption"
ISSUE_STRUCT   = "structural_gap"
ISSUE_CATALOG  = "catalog_verify"


# ── helpers ───────────────────────────────────────────────────────────────────

def load_programs(programs_dir):
    programs = []
    for path in sorted(glob.glob(os.path.join(programs_dir, "*.json"))):
        with open(path) as f:
            data = json.load(f)
        data["_path"] = path
        programs.append(data)
    programs.sort(key=lambda d: (d["program"]["college"], d["program"]["id"]))
    return programs


def iter_all_groups(data):
    """Yield every group dict from major (groups or phases) and gen_ed tracks."""
    major = data.get("major", {})
    if "groups" in major:
        yield from major["groups"]
    if "phases" in major:
        for phase in major["phases"]:
            yield from phase.get("groups", [])
    for track in data.get("general_education", {}).get("tracks", []):
        yield from track.get("groups", [])


def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def apply_row_style(ws, row, ncols, alt=False, link_col=None):
    fill = ALT_FILL if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        if link_col and col == link_col:
            cell.font = LINK_FONT
        else:
            cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=(col >= ncols - 1))
        cell.border = THIN_BORDER


# ── Sheet 1: Program Review ───────────────────────────────────────────────────

SHEET1_HEADERS = [
    "Program ID", "Title", "Sequence", "Degree",
    "College", "Catalog", "Reviewer", "Status", "Notes",
]
SHEET1_WIDTHS = [36, 22, 30, 8, 32, 12, 18, 18, 40]

STATUS_OPTIONS = ["Pending", "Approved", "Changes Needed"]


def build_sheet1(wb, programs):
    ws = wb.create_sheet("Program Review")

    ws.append(SHEET1_HEADERS)
    style_header_row(ws, len(SHEET1_HEADERS))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    dv_status = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_status)

    for i, data in enumerate(programs, 2):
        prog = data["program"]
        pid  = prog["id"]
        url  = BASE_URL + pid + ".html"
        cat  = prog.get("catalog_url", "")

        ws.cell(i, 1, pid).hyperlink  = url
        ws.cell(i, 2, prog.get("title", ""))
        ws.cell(i, 3, prog.get("sequence", ""))
        ws.cell(i, 4, prog.get("degree", ""))
        ws.cell(i, 5, prog.get("college", ""))
        if cat:
            ws.cell(i, 6, "Catalog").hyperlink = cat
        ws.cell(i, 7, "")   # Reviewer
        ws.cell(i, 8, "Pending")
        ws.cell(i, 9, "")   # Notes

        dv_status.add(ws.cell(i, 8))
        apply_row_style(ws, i, len(SHEET1_HEADERS), alt=(i % 2 == 0), link_col=1)

        # Style catalog link cell separately
        cat_cell = ws.cell(i, 6)
        if cat:
            cat_cell.font = LINK_FONT
        else:
            cat_cell.font = NORMAL_FONT

    set_col_widths(ws, SHEET1_WIDTHS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SHEET1_HEADERS))}1"
    return ws


# ── Sheet 2: Issue Log ────────────────────────────────────────────────────────

SHEET2_HEADERS = [
    "Program ID", "College", "Issue Type", "Group ID",
    "Description", "Resolution", "Resolved",
]
SHEET2_WIDTHS = [36, 32, 18, 36, 60, 40, 10]

RESOLVED_OPTIONS = ["No", "Yes", "N/A"]

# Manually-known structural gaps that need advisor attention
STRUCTURAL_GAPS = [
    {
        "program_id": "accntcybs-financial-accounting",
        "college": "College of Business",
        "issue_type": ISSUE_STRUCT,
        "group_id": "major.required_courses",
        "description": (
            "Required block contains inline choose_one slots for math, writing, and IT options "
            "per the catalog, but is currently encoded as a flat fixed list. "
            "Advisor: identify which slots are choices and confirm the options for each."
        ),
    },
    {
        "program_id": "artba-art-history",
        "college": "Wonsook Kim Coll of Fine Arts",
        "issue_type": ISSUE_CATALOG,
        "group_id": "major.required_courses",
        "description": (
            "Language sequence (FRE/GER/ITA/SPA 111/112/115) is encoded as three inline "
            "choose_one slots. Advisor: confirm whether students must complete all three "
            "levels in the same language (i.e., pick French and stay in French) or whether "
            "they may mix languages across the sequence."
        ),
    },
    {
        "program_id": "artba-art-history",
        "college": "Wonsook Kim Coll of Fine Arts",
        "issue_type": ISSUE_CREDIT,
        "group_id": "major.art_history_elective_courses",
        "description": (
            "12 elective courses (ART 240-281) assigned 3 credit hours each based on "
            "standard course format. Advisor: verify actual credit values and confirm "
            "whether any per-group minimum applies (catalog lists three groups of courses "
            "but does not explicitly state a distribution requirement)."
        ),
    },
    {
        "program_id": "nurbsn-r-n-to-b-s-n",
        "college": "Mennonite College of Nursing",
        "issue_type": ISSUE_CREDIT,
        "group_id": "major.escrow",
        "description": (
            "Catalog note states 34 total escrow credits (NUR 229/231/314/316/317/325), "
            "but individual course credits from legacy data sum to 32. "
            "Advisor: confirm total escrow credit value with the college."
        ),
    },
    {
        "program_id": "tchecebs-pedagogy",
        "college": "College of Education",
        "issue_type": ISSUE_STRUCT,
        "group_id": "major.complete_one_of_the_following (x9)",
        "description": (
            "Nine separate choose_n(1) groups may represent teaching-specialty track bundles "
            "rather than independent choices. Advisor: confirm whether students choose one "
            "specialty track (bundle of courses) or make nine independent course selections."
        ),
    },
    {
        "program_id": "musbm-composition-theory-emphasis",
        "college": "Wonsook Kim Coll of Fine Arts",
        "issue_type": ISSUE_STRUCT,
        "group_id": "major.applied_music / major.performance_ensembles",
        "description": (
            "Applied music and performance ensembles are encoded as open groups with course "
            "options listed in free-text notes. Advisor: confirm credit range per enrollment, "
            "total hours required, and whether a level-progression rule applies "
            "(e.g., must take 231 before 331)."
        ),
    },
    {
        "program_id": "nurbsn-traditional-prelicensure",
        "college": "Mennonite College of Nursing",
        "issue_type": ISSUE_STRUCT,
        "group_id": "major.required_courses",
        "description": (
            "38 courses are in a flat required list. The catalog describes a phased structure "
            "(pre-nursing foundation, nursing core, clinical). Advisor: identify which courses "
            "belong to each phase so the JSON can be split into phases[]."
        ),
    },
    {
        "program_id": "accntcybs-financial-accounting",
        "college": "College of Business",
        "issue_type": ISSUE_CATALOG,
        "group_id": "isu.mathematics / iai.mathematics",
        "description": (
            "MAT 121 carries CourseDog attribute GE14-QR (Quantitative Reasoning), not "
            "GE14-MAT (Mathematics). Advisor: confirm whether MAT 121 satisfies the "
            "Mathematics gen-ed requirement or only presupposes it."
        ),
    },
]


def is_afb_verified(group, course_fulfills):
    """Return True if every course in auto_fulfilled_by is confirmed by fulfills data.

    A course is confirmed when its fulfills list includes both a major.* group
    (affirmatively required by the major) and the gen-ed group ID (affirmatively
    meets that gen-ed requirement per CourseDog attributes).
    """
    gid = group.get("id", "")
    for cid in group.get("auto_fulfilled_by", []):
        f = course_fulfills.get(cid, set())
        has_major = any(x.startswith("major.") for x in f)
        has_gened = gid in f
        if not (has_major and has_gened):
            return False
    return True


def collect_auto_fulfilled_issues(programs):
    """Return issue rows only for auto_fulfilled_by groups not verified by fulfills data."""
    issues = []
    skipped = 0

    for data in programs:
        pid     = data["program"]["id"]
        college = data["program"].get("college", "")
        course_fulfills = {c["id"]: set(c.get("fulfills", [])) for c in data.get("courses", [])}

        for group in iter_all_groups(data):
            afb = group.get("auto_fulfilled_by", [])
            if not afb:
                continue
            if is_afb_verified(group, course_fulfills):
                skipped += 1
                continue
            courses_str = ", ".join(afb)
            issues.append({
                "program_id":  pid,
                "college":     college,
                "issue_type":  ISSUE_AUTO,
                "group_id":    group.get("id", ""),
                "description": (
                    f"auto_fulfilled_by: {courses_str}. "
                    f"Verify whether these courses satisfy or merely presuppose '{group.get('id', '')}'. "
                    "If they satisfy it, the annotation is correct. "
                    "If they only presuppose it (student must still take the gen-ed separately), "
                    "remove auto_fulfilled_by."
                ),
            })

    return issues, skipped


def build_sheet2(wb, programs):
    ws = wb.create_sheet("Issue Log")

    ws.append(SHEET2_HEADERS)
    style_header_row(ws, len(SHEET2_HEADERS))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    dv_resolved = DataValidation(
        type="list",
        formula1='"' + ",".join(RESOLVED_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_resolved)

    auto_issues, skipped = collect_auto_fulfilled_issues(programs)
    all_issues = STRUCTURAL_GAPS + auto_issues

    for i, issue in enumerate(all_issues, 2):
        ws.cell(i, 1, issue["program_id"])
        ws.cell(i, 2, issue["college"])
        ws.cell(i, 3, issue["issue_type"])
        ws.cell(i, 4, issue["group_id"])
        ws.cell(i, 5, issue["description"])
        ws.cell(i, 6, "")   # Resolution
        ws.cell(i, 7, "No")

        dv_resolved.add(ws.cell(i, 7))
        apply_row_style(ws, i, len(SHEET2_HEADERS), alt=(i % 2 == 0))

    set_col_widths(ws, SHEET2_WIDTHS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SHEET2_HEADERS))}1"

    # Row heights — description column wraps
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 42

    return ws, skipped


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate advisor review workbook")
    parser.add_argument("--out", default="output/review/advisor-review.xlsx")
    parser.add_argument("--programs-dir", default=PROGRAMS_DIR)
    args = parser.parse_args()

    programs = load_programs(args.programs_dir)
    print(f"Loaded {len(programs)} programs")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_sheet1(wb, programs)
    _, skipped = build_sheet2(wb, programs)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")

    auto_total = sum(
        1
        for data in programs
        for group in iter_all_groups(data)
        if group.get("auto_fulfilled_by")
    )
    auto_flagged = auto_total - skipped
    print(f"  Sheet 1: {len(programs)} programs")
    print(f"  Sheet 2: {len(STRUCTURAL_GAPS)} structural gaps + {auto_flagged} unverified auto_fulfilled_by"
          f" ({skipped} verified by fulfills data, omitted)")


if __name__ == "__main__":
    main()
